import openpyxl
import requests
import telebot
import time
import datetime
import sqlite3
from telebot import types
from datetime import date


bot_token = '6224714194:AAERfVMEJn87MunWtYcH8MAtAOkRVelPynE'
bot = telebot.TeleBot(bot_token)

excel_file_path = 'D:\kmb2-1.xlsx'

today = date.today()
weekday_number = today.weekday()

# def time():
#     while True:
#         current_time = datetime.datetime.now().strftime('%H:%M')
#         return current_time
#
#
# time = time()


conn = sqlite3.connect('kmb2111.db', check_same_thread=False)
cursor = conn.cursor()
cursor.execute('''CREATE TABLE IF NOT EXISTS chat_ids
                (id INTEGER PRIMARY KEY)''')

conn2 = sqlite3.connect('kmb2112.db', check_same_thread=False)
cursor2 = conn2.cursor()
cursor2.execute('''CREATE TABLE IF NOT EXISTS chat_ids
                (id INTEGER PRIMARY KEY)''')

@bot.message_handler(commands=['start'])
def start_command(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)

    item1 = types.KeyboardButton("КМБ-21-1")
    item2 = types.KeyboardButton("КМБ-22-1")
    item3 = types.KeyboardButton("ИТС-21-1")
    markup.add(item1, item2, item3)

    bot.send_message(message.chat.id, "Привет! Нажми на кнопку, чтобы выбрать свою академическую группу", reply_markup=markup)


@bot.message_handler(content_types='text')
def message_reply(message):
    markup2 = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup3 = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup4 = types.ReplyKeyboardMarkup(resize_keyboard=True)

    item4 = types.KeyboardButton(text="КМБ-21-1(1)")
    item5 = types.KeyboardButton(text="КМБ-21-1(2)")
    markup2.add(item4, item5)

    item6 = types.KeyboardButton("КМБ-22-1(1)")
    item7 = types.KeyboardButton("КМБ-22-1(2)")
    markup3.add(item6, item7)

    item8 = types.KeyboardButton("ИТС-21-1(1)")
    item9 = types.KeyboardButton("ИТС-21-1(2)")
    markup4.add(item8, item9)

    if message.text == "КМБ-21-1":
        bot.send_message(message.chat.id, "Выберите подгруппу", reply_markup=markup2)
    if message.text == "КМБ-21-1(1)":
        bot.send_message(message.chat.id, "Теперь вам будет приходить ваше расписание", reply_markup=types.ReplyKeyboardRemove())
        chat_id = message.chat.id
        # проверяем, есть ли уже ID чата в базе данных
        cursor.execute('SELECT * FROM chat_ids WHERE id = ?', (chat_id,))
        result = cursor.fetchone()
        if result is None:
            cursor.execute('INSERT INTO chat_ids VALUES (?)', (chat_id,))
            conn.commit()
            # bot.send_message(message.chat.id, 'ID чата сохранен в базе данных')
        # else:
            # bot.send_message(message.chat.id, 'ID чата уже существует в базе данных')
        if result:
            if weekday_number == 6:
                wb = openpyxl.load_workbook(excel_file_path)
                ws = wb.active
                data = ws['A1'].value
                bot.send_message(message.chat.id, data)


    if message.text == "КМБ-21-1(2)":
        bot.send_message(message.chat.id, "Теперь вам будет приходить ваше расписание", reply_markup=types.ReplyKeyboardRemove())
        chat_id = message.chat.id
        # проверяем, есть ли уже ID чата в базе данных
        cursor2.execute('SELECT * FROM chat_ids WHERE id = ?', (chat_id,))
        result = cursor2.fetchone()
        if result is None:
            cursor2.execute('INSERT INTO chat_ids VALUES (?)', (chat_id,))
            conn2.commit()
            bot.send_message(message.chat.id, 'ID чата сохранен в базе данных')
        else:
            bot.send_message(message.chat.id, 'ID чата уже существует в базе данных')
        # if weekday_number == 6:
        #     if time == "18:27":
        #         wb = openpyxl.load_workbook(excel_file_path)
        #         ws = wb.active
        #         data = ws['A1'].value
        #         bot.send_message(message.chat.id, data)


    if message.text == "КМБ-22-1":
        bot.send_message(message.chat.id, "Выберите подгруппу", reply_markup=markup3)


    if message.text == "ИТС-21-1":
        bot.send_message(message.chat.id, "Выберите подгруппу", reply_markup=markup4)

bot.infinity_polling()
